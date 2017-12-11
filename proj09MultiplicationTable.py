#! /usr/bin/python3
# proj09MultiplicationTable.py - Creates a multiplication table of size n.
# https://www.reddit.com/r/beginnerprojects/comments/2agwnq/project_multiplication_table/
# Adapted from a practice project in How to Automate the Boring Stuff with Python, by Al Sweigert. It stores table in a .xlsx file rather than printing to screen.

import openpyxl, sys
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Prompt user for size of table.
N = int(input("What size table would you like?"))

# Create new workbook
wb = openpyxl.Workbook()
sheet = wb.active

# Create bold, frozen pane headers in row 1 & column 1.
sheet.freeze_panes = 'B2'
boldFont = Font(bold=True)
for i in range(2, N+2):
    sheet['A' + str(i)].font = boldFont
    sheet['A' + str(i)].value = i-1
for i in range(2, N+2):
    sheet[get_column_letter(i) + '1'].font = boldFont
    sheet[get_column_letter(i) + '1'].value = i-1

# Populate multiplication table
for i in range(2, N+2):
    for j in range(2, N+2):
        sheet[get_column_letter(i) + str(j)] = (i-1)*(j-1)

# Save file
wb.save('multTable' + str(N) + '.xlsx')
